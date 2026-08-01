"""Sales ingestion — parse an upload into canonical rows, then upsert (venue-scoped).

Two seams keep the core importable and vendor-neutral:

* **parsing** is generic. A CSV (stdlib ``csv``) or ``.xlsx`` (lazy ``openpyxl``,
  imported only inside the parse call) file is read into header-keyed dicts and mapped
  to a small canonical schema via :data:`CANONICAL_FIELDS`. Vendor-specific column
  quirks (fixed 25-column exports, sheet names, currency-string cells beyond the basic
  strip) are DEFERRED — a real overlay supplies its own column map / parser.
* **storage** of the archived upload is via the injected :class:`Storage` seam.

Idempotency: each row's ``row_hash`` (sha256 over stable natural fields) is the key —
re-uploading the same bytes inserts nothing (counted as skipped). The processed
``analytics_sales`` row is enriched with a generic ``components`` breakdown from the
dish catalog (``analytics_menu``) when the (name, price) resolves; otherwise ``[]``.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
from datetime import datetime
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from be.adapters.storage.base import Storage
from be.app.domains.analytics import crud
from be.app.domains.analytics.timeutil import dow_sun0, parse_ts

# Canonical fields the importer understands. A source header (case-insensitive) that
# matches one of these names is mapped through; everything else is ignored.
CANONICAL_FIELDS: tuple[str, ...] = (
    "order_id", "order_date", "order_ts", "dish_type", "dish_name", "menu_price",
    "discount", "sale_price", "qty", "waiter_name", "diners", "cost", "cost_sum",
    "components",
)

_CURRENCY_STRIP = str.maketrans({"₪": None, ",": None, "\xa0": None})


def _to_num(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).translate(_CURRENCY_STRIP).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(value: Any) -> int | None:
    num = _to_num(value)
    return int(num) if num is not None else None


def _to_text(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _to_date(value: Any) -> str | None:
    s = _to_text(value)
    if s is None:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _parse_components(value: Any) -> list[dict]:
    if value is None or value == "":
        return []
    raw: Any = value
    if isinstance(value, str):
        try:
            raw = json.loads(value)
        except json.JSONDecodeError:
            return []
    arr = raw.get("components") if isinstance(raw, dict) else raw
    if not isinstance(arr, list):
        return []
    out = []
    for item in arr:
        if not isinstance(item, dict):
            continue
        name = item.get("name")
        try:
            count = int(str(item.get("count")))
        except (TypeError, ValueError):
            count = 0
        if name:
            out.append({"name": str(name), "count": count})
    return out


def parse_bytes(data: bytes, *, filename: str | None) -> list[dict]:
    """Parse an upload into a list of raw header-keyed dicts (CSV or ``.xlsx``)."""
    name = (filename or "").lower()
    is_xlsx = name.endswith(".xlsx") or (not name.endswith(".csv") and data[:2] == b"PK")
    if is_xlsx:
        return _parse_xlsx(data)
    text_data = data.decode("utf-8-sig", "ignore")
    reader = csv.DictReader(io.StringIO(text_data))
    rows = []
    for rec in reader:
        rows.append({(k or "").strip().lower(): v for k, v in rec.items()})
    return rows


def _parse_xlsx(data: bytes) -> list[dict]:
    # Lazy import: keeps `import be.app.domains.analytics` cheap and importable even
    # if openpyxl is not installed.
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip().lower() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    out = []
    for row in rows_iter:
        out.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
    return out


def _canonicalize(rec: dict) -> dict:
    """Map a header-keyed source row to the canonical row + derived heatmap keys."""
    row: dict[str, Any] = {f: rec.get(f) for f in CANONICAL_FIELDS}
    order_id = _to_int(row.get("order_id"))
    order_date = _to_date(row.get("order_date"))
    ts = parse_ts(_to_text(row.get("order_ts")))
    if ts is None and order_date:
        ts = parse_ts(order_date)
    canonical = {
        "order_id": order_id,
        "order_date": order_date or (ts.date().isoformat() if ts else None),
        "order_ts": ts.isoformat() if ts else None,
        "order_dow": dow_sun0(ts) if ts else None,
        "order_hour_int": ts.hour if ts else None,
        "dish_type": _to_text(row.get("dish_type")),
        "dish_name": _to_text(row.get("dish_name")),
        "menu_price": _to_num(row.get("menu_price")),
        "discount": _to_num(row.get("discount")),
        "sale_price": _to_num(row.get("sale_price")),
        "qty": _to_num(row.get("qty")),
        "waiter_name": _to_text(row.get("waiter_name")),
        "diners": _to_int(row.get("diners")),
        "cost": _to_num(row.get("cost")),
        "cost_sum": _to_num(row.get("cost_sum")),
        "components": _parse_components(row.get("components")),
    }
    return canonical


def _row_hash(row: dict) -> str:
    def part(v: Any) -> str:
        return "" if v is None else str(v)

    parts = [
        part(row["order_id"]), part(row["order_ts"]), part(row["dish_name"]),
        part(row["menu_price"]), part(row["sale_price"]), part(row["discount"]),
        part(row["qty"]),
    ]
    return hashlib.sha256("\x1f".join(parts).encode("utf-8")).hexdigest()


async def ingest_bytes(
    session: AsyncSession,
    storage: Storage,
    *,
    venue_id: str,
    data: bytes,
    filename: str | None,
    source: str,
) -> dict:
    """Parse + upsert an upload; archive it via Storage; write an audit row.

    Returns ``{rows_inserted, rows_skipped, date_min, date_max, ingestion_id}``.
    A parse/DB error is recorded on the ingestion row and returned (never raised).
    """
    rows_inserted = 0
    rows_skipped = 0
    date_min: str | None = None
    date_max: str | None = None
    inserted_raw_ids: list[int] = []

    try:
        records = parse_bytes(data, filename=filename)
        menu = await crud.menu_lookup(session, venue_id=venue_id)
        for rec in records:
            row = _canonicalize(rec)
            if row["order_id"] is None:
                rows_skipped += 1
                continue
            row["row_hash"] = _row_hash(row)
            if await crud.row_hash_exists(
                session, table="analytics_sales_raw", row_hash=row["row_hash"]
            ):
                rows_skipped += 1
                continue

            raw_id = await crud.insert_raw_row(
                session, venue_id=venue_id, row=row, source=source
            )
            inserted_raw_ids.append(raw_id)

            components = row["components"]
            if not components:
                m = menu.get((row["dish_name"], float(row["menu_price"] or 0)))
                if m:
                    components = m["components"]
            await crud.insert_sales_row(
                session, venue_id=venue_id, row=row, source=source,
                raw_id=raw_id, components=components,
            )

            rows_inserted += 1
            od = row["order_date"]
            if od:
                date_min = od if date_min is None else min(date_min, od)
                date_max = od if date_max is None else max(date_max, od)
        await session.commit()
    except Exception as exc:  # noqa: BLE001 - record failure in the audit log
        await session.rollback()
        ingestion_id = await crud.write_ingestion(
            session, venue_id=venue_id, source=source, filename=filename,
            rows_inserted=0, rows_skipped=0, date_min=None, date_max=None,
            status="error", error=str(exc)[:2000], file_path=None,
        )
        await session.commit()
        return {
            "rows_inserted": 0, "rows_skipped": 0, "date_min": None, "date_max": None,
            "ingestion_id": ingestion_id, "error": str(exc),
        }

    ingestion_id = await crud.write_ingestion(
        session, venue_id=venue_id, source=source, filename=filename,
        rows_inserted=rows_inserted, rows_skipped=rows_skipped,
        date_min=date_min, date_max=date_max, status="success", error=None,
        file_path=None,
    )
    if inserted_raw_ids:
        await crud.tag_raw_ingestion(
            session, ids=inserted_raw_ids, ingestion_id=ingestion_id
        )
    await session.commit()

    # Best-effort archive via the Storage seam; a failure never fails the ingest.
    file_path = _archive_key(data, filename, ingestion_id)
    try:
        await storage.put(file_path, data, _content_type(filename))
        await crud.set_ingestion_file(session, ing_id=ingestion_id, file_path=file_path)
        await session.commit()
    except Exception:  # noqa: BLE001 - archival is optional
        await session.rollback()

    await crud.log_processing_run(
        session, venue_id=venue_id, kind="ingest",
        params={"ingestion_id": ingestion_id, "source": source, "filename": filename},
        rows_affected=rows_inserted, status="success", actor=source,
    )
    await session.commit()

    return {
        "rows_inserted": rows_inserted, "rows_skipped": rows_skipped,
        "date_min": date_min, "date_max": date_max, "ingestion_id": ingestion_id,
    }


def _archive_key(data: bytes, filename: str | None, ingestion_id: int) -> str:
    digest = hashlib.sha256(data).hexdigest()[:12]
    safe = (filename or "upload").replace("/", "_").replace(" ", "_")
    return f"sales_uploads/{ingestion_id}_{digest}_{safe}"


def _content_type(filename: str | None) -> str:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return "text/csv"
    if name.endswith(".xlsx"):
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return "application/octet-stream"


async def reprocess(
    session: AsyncSession, *, venue_id: str, ingestion_id: int | None,
    date_from: str | None, date_to: str | None, actor: str | None,
) -> dict:
    """Rebuild processed sales from raw for one upload or a date range (portable)."""
    try:
        deleted = await crud.delete_sales_for_reprocess(
            session, venue_id=venue_id, ingestion_id=ingestion_id,
            date_from=date_from, date_to=date_to,
        )
        raw_rows = await crud.raw_rows_for_reprocess(
            session, venue_id=venue_id, ingestion_id=ingestion_id,
            date_from=date_from, date_to=date_to,
        )
        menu = await crud.menu_lookup(session, venue_id=venue_id)
        inserted = 0
        for raw in raw_rows:
            row = dict(raw)
            row["order_id"] = raw["order_id"]
            components = crud.load_json(raw["components"])
            if not components:
                m = menu.get((raw["dish_name"], float(raw["menu_price"] or 0)))
                if m:
                    components = m["components"]
            if await crud.row_hash_exists(
                session, table="analytics_sales", row_hash=raw["row_hash"]
            ):
                continue
            await crud.insert_sales_row(
                session, venue_id=venue_id, row=row, source=raw["source"],
                raw_id=raw["id"], components=components,
            )
            inserted += 1
        await crud.log_processing_run(
            session, venue_id=venue_id, kind="reprocess",
            params={"ingestion_id": ingestion_id, "date_from": date_from,
                    "date_to": date_to, "deleted": deleted, "inserted": inserted},
            rows_affected=inserted, status="success", actor=actor,
        )
        await session.commit()
    except Exception as exc:  # noqa: BLE001
        await session.rollback()
        await crud.log_processing_run(
            session, venue_id=venue_id, kind="reprocess",
            params={"ingestion_id": ingestion_id}, rows_affected=0,
            status="error", error=str(exc)[:2000], actor=actor,
        )
        await session.commit()
        raise
    return {"deleted": deleted, "inserted": inserted}


__all__ = ["CANONICAL_FIELDS", "parse_bytes", "ingest_bytes", "reprocess"]
